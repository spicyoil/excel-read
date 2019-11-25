import openpyxl
import pprint
import os
#print(os.getcwd())
f = open("./reg_mode_check.txt", 'w+')
files = os.listdir(".")

for eachfile in files:
    if (".xlsm" in eachfile):
    #if eachfile=="ISX020_CUTOP_CUSTOM26_レジスタマップ.xlsm":
        print("Processing " + eachfile)
        wb = openpyxl.load_workbook(eachfile)

        for sheetname in wb.sheetnames:
            if (('マクロ凡例' not in sheetname) and ('設定' not in sheetname) and ('変更履歴' not in sheetname) and ('sheet' not in sheetname) and ('Sheet' not in sheetname)):
                sheet= wb[sheetname]
                thisheetname = sheetname
                #print("Proceesing " + thisheetname)


        # print(sheet["H"])    # (<Cell Sheet3.C1>, <Cell Sheet3.C2>, <Cell Sheet3.C3>, <Cell Sheet3.C4>, <Cell Sheet3.C5>, <Cell Sheet3.C6>, <Cell Sheet3.C7>, <Cell Sheet3.C8>, <Cell Sheet3.C9>, <Cell Sheet3.C10>)      <-第C列
        # print(sheet["4"])    # (<Cell Sheet3.A4>, <Cell Sheet3.B4>, <Cell Sheet3.C4>, <Cell Sheet3.D4>, <Cell Sheet3.E4>)     <-第4行
        # print(sheet["H4"].value)  #c4     <-第C4格的值
        # print(sheet.rows)         #<generator object _cells_by_row at 0x0000000003A2EB40>             获取表格所有行
        # print(sheet.columns)      #<generator object _cells_by_col at 0x0000000003A2EB40>             获取表格所有列
        # print(sheet.max_row)      # 10     <-最大行数
        # print(sheet.max_column)   # 5     <-最大列数
        # for i in sheet["H"]:
        #     print(i.value, end=" ")  # c1 c2 c3 c4 c5 c6 c7 c8 c9 c10     <-C列中的所有值
        # for i in sheet['F3:H5']:
        #     for j in i :
        #         print(j.value, end=" ")  # c1 c2 c3 c4 c5 c6 c7 c8 c9 c10     <-C列中的所有值
        #d = sheet.cell(row=2, column=1).value #read cell
        #d = ws['A4'] # read cell 

        cnt = 1
        sheet_FLAG = 1 # this sheet have no target sheet_FLAG
        regname = 0 # reg name save 
        for row in sheet.rows:
            if row[17].value != "CONST": #only CONST will be count
                continue
            
            cnt = cnt + 1
            if cnt > 2 : #exclude title
                col_index = 52 
                for i in row[52:67]: #counting area
                    if (i.value != None) and (i.value != ""):
                        #print("register: {} i.value: {} Mode {}: {}".format(row[7].value,i.value, sheet.cell(row=2,column=col_index+1).value, row[col_index].value))
                        #print(i.value,col_index,row[col_index].value)
                        #print (i, row[65])
                        # if  (i == row[65]):
                        #     continue
                        #     print ("power")

                        if sheet_FLAG :
                            print("\n"+eachfile,file=f)
                            print(thisheetname,file=f)
                            sheet_FLAG = 0

                        if  regname != row[7].value:
                            regname = row[7].value
                            print("             {}  HW: {}  Init: {} ".format(regname, row[12].value, row[13].value),file=f)
                        
                        print("             Mode({}) : {}".format(sheet.cell(row=2,column=col_index+1).value, row[col_index].value),file=f)
                    col_index = col_index + 1

print ("Wrote in reg_mode_check.txt successfully" )
